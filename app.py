# app.py
import os
import time
import base64
import requests
import sqlite3
import json
from flask import Flask, request, jsonify, render_template, session, redirect, url_for, flash, Response
from threading import Lock, Timer
from openpyxl import load_workbook
from dotenv import load_dotenv
from functools import wraps
from datetime import datetime, timedelta
from utils.task_queue import add_task
import queue
import logging
from logging.handlers import RotatingFileHandler
# ---------------- Flask Initialization ----------------
load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET", "supersecretkey")

# === Rotating File Logs Setup ===
if not os.path.exists('logs'):
    os.makedirs('logs')

log_file = 'logs/app.log'
file_handler = RotatingFileHandler(log_file, maxBytes=5 * 1024 * 1024, backupCount=10)
file_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Attach logger to Flask app
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)
# ---------------- Config / Paths ----------------
DB_PATH = os.path.join(os.getcwd(), "zoom_callerid.db")
REPORT_FILE = os.path.join(os.getcwd(), "update_report.json")

# canonical base URL name
BASE_URL = os.getenv("ZOOM_BASE_URL", "https://api.zoom.us/v2")
# alias for older code that used ZOOM_BASE_URL
ZOOM_BASE_URL = BASE_URL

# Zoom credentials (read from .env)
CLIENT_ID = os.getenv("ZOOM_CLIENT_ID")
CLIENT_SECRET = os.getenv("ZOOM_CLIENT_SECRET")
# Accept either env var name ZOOM_ACCOUNT_ID or ACCOUNT_ID if you used either
ACCOUNT_ID = os.getenv("ZOOM_ACCOUNT_ID") or os.getenv("ACCOUNT_ID")

# ---------------- Token Cache ----------------
token_cache = {"access_token": None, "expiry": 0}
token_lock = Lock()

# ---------------- In-memory report cache (kept in sync with DB) ----------------
report_cache = {"updates": []}
if os.path.exists(REPORT_FILE):
    try:
        with open(REPORT_FILE, "r", encoding="utf-8") as f:
            report_cache["updates"] = json.load(f)
    except Exception as e:
        print("Warning: couldn't load report cache:", e)
        report_cache["updates"] = []

def save_report_json():
    try:
        with open(REPORT_FILE, "w", encoding="utf-8") as f:
            json.dump(report_cache["updates"], f, indent=2, default=str)
    except Exception as e:
        print("Warning: couldn't save report json:", e)

# ---------------- DB Init & helpers ----------------
def init_db():
    """Create DB and required tables (safe to call multiple times)."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # users login table
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            email TEXT PRIMARY KEY,
            password TEXT,
            role TEXT
        )
    ''')

    # caller ID update logs
    c.execute('''
        CREATE TABLE IF NOT EXISTS update_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT,
            identifier TEXT,
            caller_id TEXT,
            alias TEXT,
            success INTEGER,
            reason TEXT,
            type TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    # 🔹 Ensure alias column exists for backward compatibility
    try:
        c.execute("PRAGMA table_info(update_logs)")
        columns = [row[1] for row in c.fetchall()]
        if "alias" not in columns:
            c.execute("ALTER TABLE update_logs ADD COLUMN alias TEXT")
            print("✅ Added missing 'alias' column to update_logs table")
    except Exception as e:
        print("⚠️ Error checking or adding alias column:", e)

    # admin activity logs
    c.execute('''
        CREATE TABLE IF NOT EXISTS admin_activity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_email TEXT,
            action TEXT,
            target_email TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # login activity to track online/offline
    c.execute('''
        CREATE TABLE IF NOT EXISTS login_activity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_email TEXT,
            login_time DATETIME,
            logout_time DATETIME,
            status TEXT
        )
    ''')

    conn.commit()
    conn.close()

init_db()

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def load_users_from_db():
    conn = get_db_connection()
    rows = conn.execute("SELECT email, password, role FROM users").fetchall()
    conn.close()
    users = {r["email"].lower(): {"password": r["password"], "role": r["role"]} for r in rows}
    return users

def save_user_to_db(email, password, role):
    conn = get_db_connection()
    conn.execute("INSERT OR REPLACE INTO users (email, password, role) VALUES (?, ?, ?)",
                 (email.lower(), password, role))
    conn.commit()
    conn.close()

def delete_user_from_db(email):
    conn = get_db_connection()
    conn.execute("DELETE FROM users WHERE email = ?", (email.lower(),))
    conn.commit()
    conn.close()
# --- Online/Offline status helper ---
def get_online_status():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT u.email,
               CASE 
                    WHEN EXISTS (
                        SELECT 1 FROM login_activity la 
                        WHERE la.user_email = u.email AND la.status = 'Online'
                    ) THEN 'Online'
                    ELSE 'Offline'
               END AS status
        FROM users u
        ORDER BY u.email
    """).fetchall()
    conn.close()
    return rows

# Seed default admins if none exist
_users = load_users_from_db()
if not _users:
    save_user_to_db("sanjita.das@blackbox.com", "Admin@123", "admin")
    save_user_to_db("rajeev.gupta@blackbox.com", "Admin@123", "admin")

# ---------------- Token generation (client_credentials) ----------------
def generate_access_token():
    """Generate Zoom access token (account_credentials grant)"""
    with token_lock:
        if token_cache["access_token"] and time.time() < token_cache["expiry"]:
            return token_cache["access_token"]

        auth_header = base64.b64encode(f"{CLIENT_ID}:{CLIENT_SECRET}".encode()).decode()
        url = f"https://zoom.us/oauth/token?grant_type=account_credentials&account_id={ACCOUNT_ID}"
        headers = {"Authorization": f"Basic {auth_header}"}

        response = requests.post(url, headers=headers)
        response.raise_for_status()
        data = response.json()

        token_cache["access_token"] = data["access_token"]
        token_cache["expiry"] = time.time() + data["expires_in"] - 60  # refresh 1 min early
        return token_cache["access_token"]

def get_zoom_headers():
    token = generate_access_token()
    return {"Authorization": f"Bearer {token}", "Accept": "application/json"}
# ---------------- Zoom API wrappers ----------------
def zoom_get_users():
    """Fetch Zoom Phone users"""
    headers = get_zoom_headers()
    url = f"{ZOOM_BASE_URL}/phone/users?page_size=100"
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    return resp.json().get("users", [])

def get_user_extension_id(email):
    """Get Zoom user extension_id by email"""
    users_list = zoom_get_users()
    for u in users_list:
        if u.get("email", "").lower() == email.lower():
            return u.get("extension_id")
    return None

def get_line_keys(extension_id):
    """Fetch line keys for a given extension"""
    headers = get_zoom_headers()
    url = f"{ZOOM_BASE_URL}/phone/extension/{extension_id}/line_keys"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        return []
    return resp.json().get("line_keys", [])

def patch_line_key(extension_id, line_key_id, new_caller_id, alias=None):
    """Update outbound caller ID and alias for a line key."""
    headers = get_zoom_headers()
    url = f"{ZOOM_BASE_URL}/phone/extension/{extension_id}/line_keys"

    payload = {
        "line_keys": [
            {
                "line_key_id": line_key_id,
                "index": 1,
                "outbound_caller_id": new_caller_id
            }
        ]
    }

    # If alias provided, include it
    if alias:
        payload["line_keys"][0]["alias"] = alias

    try:
        resp = requests.patch(url, headers=headers, json=payload)
        if resp.status_code in (200, 204):
            return True, "Updated"
        else:
            return False, resp.text
    except Exception as e:
        return False, str(e)
# ---------------- Real-Time Event System (SSE) ----------------
event_queue = queue.Queue()

def notify_clients(data):
    """Send update event to connected clients."""
    event_queue.put(data)

@app.route("/events")
def stream_events():
    """Stream updates to connected clients (SSE)."""
    def event_stream():
        while True:
            data = event_queue.get()
            yield f"data: {json.dumps(data)}\n\n"
    return Response(event_stream(), mimetype="text/event-stream")
# ---------------- Logging helpers (DB + JSON) ----------------
def log_update_db(email, identifier, caller_id, success, reason, type_, alias=None):
    conn = get_db_connection()
    conn.execute("""
        INSERT INTO update_logs (email, identifier, caller_id, alias, success, reason, type, timestamp)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (email, identifier, caller_id, alias, int(bool(success)), reason, type_, datetime.utcnow().isoformat()))
    conn.commit()
    conn.close()

def log_admin_action_db(admin_email, action, target_email):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO admin_activity (admin_email, action, target_email, timestamp) VALUES (?, ?, ?, ?)",
        (admin_email, action, target_email, datetime.utcnow().isoformat())
    )
    conn.commit()
    conn.close()

def append_report_cache(entry):
    # keep cache ordered newest first (prepend)
    report_cache["updates"].insert(0, entry)
    # keep a reasonable limit in JSON cache
    report_cache["updates"] = report_cache["updates"][:5000]
    save_report_json()

def log_update(identifier, caller_id, success, reason, type_, updated_by, alias=None):
    log_update_db(updated_by, identifier, caller_id, success, reason, type_, alias)
    entry = {
        "email": updated_by,
        "identifier": identifier,
        "caller_id": caller_id,
        "alias": alias or "-",
        "success": bool(success),
        "reason": reason,
        "type": type_,
        "time": datetime.utcnow().strftime("%d-%b-%Y %H:%M:%S")
    }
    append_report_cache(entry)
    notify_clients(entry)
# ---------------- Background Task Worker Functions ----------------
def process_single_update_task(identifier, caller_id, alias, update_type, user_email):
    """This runs in background when single update is queued."""
    print(f"[Task] Starting queued single update for {identifier}")

    ext_id = None
    email_updated = None

    if update_type == "email":
        email_updated = identifier
        ext_id = get_user_extension_id(email_updated)

    elif update_type == "extension_number":
        ext_id = get_extension_id_from_number(identifier)
        email_updated = get_email_from_extension_id(ext_id) or f"ExtNum_{identifier}"

    elif update_type == "extension_id":
        ext_id = identifier
        email_updated = get_email_from_extension_id(ext_id) or f"Extension_{ext_id}"

    else:
        log_update(identifier, caller_id, False, "Invalid update type", "S", user_email)
        return

    if not ext_id:
        log_update(identifier, caller_id, False, "No extension found", "S", user_email)
        return

    line_keys = get_line_keys(ext_id)
    if not line_keys:
        log_update(identifier, caller_id, False, "No line keys found", "S", user_email)
        return

    for lk in line_keys:
        current_id = lk.get("key_assignment", {}).get("phone_number", "")
        if current_id == caller_id:
            log_update(identifier, caller_id, False, "Caller ID already same", "S", user_email, alias)
            continue

        success, reason = patch_line_key(ext_id, lk.get("line_key_id"), caller_id, alias)
        log_update(identifier, caller_id, success, reason, "S", user_email, alias)

    print(f"[Task] Single update completed for {identifier}")


def process_bulk_update_task(rows, update_type, user_email):
    """Background worker for bulk Excel update."""
    print(f"[Task] Starting queued bulk update for {len(rows)} rows")

    for idx, (identifier, caller_id, alias) in enumerate(rows, start=1):
        try:
            ext_id = None
            email_updated = None

            if update_type == "email":
                email_updated = identifier
                ext_id = get_user_extension_id(email_updated)

            elif update_type == "extension_number":
                ext_id = get_extension_id_from_number(identifier)
                email_updated = get_email_from_extension_id(ext_id) or f"ExtNum_{identifier}"

            elif update_type == "extension_id":
                ext_id = identifier
                email_updated = get_email_from_extension_id(ext_id) or f"ExtID_{identifier}"

            else:
                log_update(identifier, caller_id, False, "Invalid update type", "B", user_email, alias)
                continue

            if not ext_id:
                log_update(identifier, caller_id, False, "No extension found", "B", user_email, alias)
                continue

            line_keys = get_line_keys(ext_id)
            if not line_keys:
                log_update(identifier, caller_id, False, "No line keys found", "B", user_email, alias)
                continue

            for lk in line_keys:
                current_id = lk.get("key_assignment", {}).get("phone_number", "")
                if current_id == caller_id:
                    log_update(identifier, caller_id, False, "Caller ID already same", "B", user_email, alias)
                    continue

                success, reason = patch_line_key(ext_id, lk.get("line_key_id"), caller_id, alias)
                log_update(identifier, caller_id, success, reason, "B", user_email, alias)

            if idx % 10 == 0:
                print(f"[Task] Processed {idx} rows...")

        except Exception as e:
            log_update(identifier, caller_id, False, f"Error: {e}", "B", user_email, alias)

    print(f"[Task] Bulk update completed ({len(rows)} records)")

# ---------------- Login activity (online/offline) ----------------
def log_login(email):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO login_activity (user_email, login_time, logout_time, status) VALUES (?, ?, ?, ?)",
        (email.lower(), datetime.utcnow().isoformat(), None, "Online")
    )
    conn.commit()
    conn.close()

def log_logout(email):
    conn = get_db_connection()
    # mark last online record as offline
    conn.execute(
        "UPDATE login_activity SET logout_time = ?, status = ? WHERE user_email = ? AND status = ?",
        (datetime.utcnow().isoformat(), "Offline", email.lower(), "Online")
    )
    conn.commit()
    conn.close()

def get_user_status_list():
    """Return list of {email, online(bool)} for users in DB (admins + users)."""
    users = load_users_from_db()
    conn = get_db_connection()
    rows = conn.execute("SELECT user_email, status, login_time FROM login_activity ORDER BY id DESC").fetchall()
    conn.close()
    status = {}
    for r in rows:
        e = r["user_email"].lower()
        if e not in status:
            status[e] = {"email": e, "online": (r["status"] == "Online"), "last_seen": r["login_time"]}
    # include users without login_activity as offline
    out = []
    for email, u in users.items():
        if email in status:
            out.append({"email": email, "online": bool(status[email]["online"])})
        else:
            out.append({"email": email, "online": False})
    return out

# ---------------- Cleanup old logs (90 days) ----------------
def cleanup_old_logs():
    try:
        cutoff = (datetime.utcnow() - timedelta(days=90)).isoformat()
        conn = get_db_connection()
        conn.execute("DELETE FROM update_logs WHERE timestamp < ?", (cutoff,))
        conn.execute("DELETE FROM admin_activity WHERE timestamp < ?", (cutoff,))
        conn.execute("DELETE FROM login_activity WHERE login_time < ?", (cutoff,))
        conn.commit()
        conn.close()
    except Exception as e:
        print("cleanup_old_logs error:", e)
    finally:
        # schedule next cleanup in 24 hours
        Timer(24 * 3600, cleanup_old_logs).start()

# start the first cleanup in background
Timer(5, cleanup_old_logs).start()  # run shortly after startup
#-------------By extension_id--------------------------------
def get_email_from_extension_id(extension_id):
    """
    Returns the email of a Zoom user given their extension_id.
    """
    users_list = zoom_get_users()  # Uses your existing zoom_get_users() function
    for user in users_list:
        if user.get("extension_id") == extension_id:
            return user.get("email")
    return None


def get_extension_id_from_number(extension_number):
    """
    Returns the extension_id from a given extension_number using Zoom API.
    """
    headers = get_zoom_headers()  # uses your token system
    url = f"{ZOOM_BASE_URL}/phone/users?page_size=100"
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
    except Exception as e:
        print("Error fetching users:", e)
        return None

    users = response.json().get("users", [])
    for user in users:
        if str(user.get("extension_number")) == str(extension_number):
            return user.get("extension_id")  # Zoom API uses extension_id
    return None

# ---------------- Login Required Decorator ----------------
def login_required(role=None):
    def wrapper(fn):
        @wraps(fn)
        def decorated(*args, **kwargs):
            if "email" not in session:
                app.logger.warning("Unauthorized access attempt to protected route.")
                return redirect(url_for("login"))
            if role and session.get("role") != role:
                app.logger.warning(f"Access denied for user {session.get('email')} - insufficient role.")
                flash("Access denied", "danger")
                return redirect(url_for("login"))
            return fn(*args, **kwargs)
        return decorated
    return wrapper


# ------------------- Login -------------------
@app.route("/", methods=["GET", "POST"])
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = (request.form.get("password") or "").strip()

        conn = get_db_connection()
        row = conn.execute(
            "SELECT * FROM users WHERE lower(email)=? AND password=?", (email, password)
        ).fetchone()
        conn.close()

        # Log every login attempt
        app.logger.info(f"Login attempt for: {email}")

        if row:
            session["email"] = row["email"]
            session["role"] = row["role"]
              
            log_login(email)  
            # Set login success flag
            session["login_success"] = True

            # Log successful login
            app.logger.info(f" Login successful for {email} (Role: {row['role']})")

            # Redirect after login
            return redirect(url_for("dashboard" if session["role"] == "admin" else "index"))
        else:
            # Log failed attempt
            app.logger.warning(f" Failed login attempt for {email}")
            flash("Invalid email or password", "danger")

    return render_template("login.html")


# ------------------- Logout -------------------
@app.route("/logout")
@login_required()
def logout():
    email = session.get("email")

    if email:
        # Log the logout in database and rotating logs
        log_logout(email)
        app.logger.info(f"User logged out: {email}")

    # Clear session
    session.clear()
    flash("Logged out successfully.", "info")

    # Redirect back to login page
    return redirect(url_for("login"))

#---------------------------------------------------dashboard--------------------------------------S
@app.route("/dashboard")
@login_required(role="admin")
def dashboard():
    conn = get_db_connection()

    # Totals
    total_admin = conn.execute("SELECT COUNT(*) as cnt FROM users WHERE role='admin'").fetchone()["cnt"]
    total_users = conn.execute("SELECT COUNT(*) as cnt FROM users WHERE role='user'").fetchone()["cnt"]

    # Last 30 days updates
    cutoff = (datetime.utcnow() - timedelta(days=30)).isoformat()
    rows = conn.execute("""
        SELECT email, identifier, caller_id, success, reason, type, timestamp 
        FROM update_logs 
        WHERE timestamp >= ? 
        ORDER BY timestamp DESC
    """, (cutoff,)).fetchall()

    updates = []
    for r in rows:
        updates.append({
            "email": r["email"],
            "identifier": r["identifier"],
            "caller_id": r["caller_id"],
            "success": bool(r["success"]),
            "reason": r["reason"],
            "type": r["type"],
            "time": datetime.fromisoformat(r["timestamp"]).strftime("%d-%b-%Y %H:%M:%S")
        })

    # Manage actions (admin_activity)
    mrows = conn.execute("""
        SELECT admin_email, action, target_email, timestamp 
        FROM admin_activity 
        WHERE timestamp >= ? 
        ORDER BY timestamp DESC
    """, (cutoff,)).fetchall()

    manage_logs = []
    for r in mrows:
        manage_logs.append({
            "email": r["admin_email"],
            "action": r["action"],
            "target": r["target_email"],
            "time": datetime.fromisoformat(r["timestamp"]).strftime("%d-%b-%Y %H:%M:%S")
        })

    conn.close()

    user_status = get_user_status_list()

    # Auto-refresh logic
    refresh_needed = session.pop("refresh_needed", False)

    return render_template(
        "dashboard.html",
        total_admin=total_admin,
        total_users=total_users,
        updates=updates,
        manage_logs=manage_logs,
        user_status=user_status,
        refresh_needed=refresh_needed
    )

@app.route("/index")
@login_required()
def index():
    # show last 30 days report
    conn = get_db_connection()
    cutoff = (datetime.utcnow() - timedelta(days=30)).isoformat()
    rows = conn.execute("SELECT identifier, caller_id, success, reason, type, timestamp FROM update_logs WHERE timestamp >= ? ORDER BY id DESC", (cutoff,)).fetchall()
    conn.close()
    report = []
    for r in rows:
        report.append({
            "email": r["identifier"],
            "caller_id": r["caller_id"],
            "success": bool(r["success"]),
            "reason": r["reason"],
            "type": r["type"],
            "time": datetime.fromisoformat(r["timestamp"]).strftime("%d-%b-%Y %H:%M:%S")
        })
    return render_template("index.html", report=report)

@app.route("/get_report")
@login_required()
def get_report():
    # returns last-30-days cached JSON (also available in DB)
    conn = get_db_connection()
    cutoff = (datetime.utcnow() - timedelta(days=30)).isoformat()
    rows = conn.execute("SELECT identifier, caller_id, success, reason, type, timestamp FROM update_logs WHERE timestamp >= ? ORDER BY id DESC", (cutoff,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        out.append({
            "identifier": r["identifier"],
            "caller_id": r["caller_id"],
            "success": bool(r["success"]),
            "reason": r["reason"],
            "type": r["type"],
            "time": datetime.fromisoformat(r["timestamp"]).strftime("%d-%b-%Y %H:%M:%S")
        })
    return jsonify({"status":"success","report":out})
#----------------------------------------------------------------single update-------------------------------------------------------------------
@app.route("/single_update", methods=["POST"])
@login_required()
def single_update():
    identifier = request.form.get("identifier", "").strip()
    caller_id = request.form.get("caller_id", "").strip()
    alias = request.form.get("alias", "").strip() or None
    update_type = request.form.get("update_type", "email").strip()

    if not identifier or not caller_id:
        return jsonify({"status": "error", "message": "Identifier and Caller ID required"}), 400

    # Queue the background task instead of processing now
    task_id = f"single_{identifier}_{int(time.time())}"
    add_task(task_id, process_single_update_task, identifier, caller_id, alias, update_type, session["email"])

    return jsonify({
        "status": "queued",
        "task_id": task_id,
        "message": f"Update for {identifier} has been queued for processing."
    })

# ---------------- Bulk Update --------------------------------------------------------------------------------------------------------
@app.route("/bulk_update", methods=["POST"])
@login_required()
def bulk_update():
    update_type = request.form.get("update_type", "email").strip()
    file = request.files.get("excel_file")
    if not file:
        return jsonify({"status": "error", "message": "Excel file required"}), 400

    wb = load_workbook(file)
    sheet = wb.active

    rows = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 2:
            continue
        if len(row) >= 3:
            identifier, caller_id, alias = row[0], row[1], row[2]
        else:
            identifier, caller_id, alias = row[0], row[1], None

        if not identifier or not caller_id:
            continue

        rows.append((str(identifier).strip(), str(caller_id).strip(), str(alias).strip() if alias else None))

    # Queue background task
    task_id = f"bulk_{int(time.time())}"
    add_task(task_id, process_bulk_update_task, rows, update_type, session["email"])

    return jsonify({
        "status": "queued",
        "task_id": task_id,
        "message": f"{len(rows)} updates queued for background processing."
    })

# ---------------- Manage Users ----------------
@app.route("/manage_users", methods=["GET","POST"])
@login_required(role="admin")
def manage_users():
    users = load_users_from_db()
    if request.method == "POST":
        action = request.form.get("action")
        email = (request.form.get("email") or "").strip().lower()
        password = (request.form.get("password") or "").strip()
        role = (request.form.get("role") or "user").strip()

        if action == "add" and email and password:
            if email in users:
                flash(f"User {email} already exists", "warning")
            else:
                save_user_to_db(email, password, role)
                log_admin_action_db(session.get("email"), "Add", email)
                flash(f"Added user {email}", "success")

        elif action == "update" and email in users:
            # allow update of password & role
            if password:
                save_user_to_db(email, password, role)
            else:
                # update role only
                save_user_to_db(email, users[email]["password"], role)
            log_admin_action_db(session.get("email"), "Update", email)
            flash(f"Updated user {email}", "success")

        elif action == "delete" and email in users:
            delete_user_from_db(email)
            log_admin_action_db(session.get("email"), "Delete", email)
            flash(f"Deleted user {email}", "success")

        users = load_users_from_db()

    sorted_users = dict(sorted(users.items()))
    return render_template("manage_users.html", users=sorted_users)

@app.route("/delete_user/<email>", methods=["POST"])
@login_required(role="admin")
def delete_user(email):
    delete_user_from_db(email)
    flash(f"Deleted user {email}", "success")
    log_admin_action_db(session.get("email"), "Delete", email)
    return redirect(url_for("manage_users"))

@app.route("/update_user_role/<email>", methods=["POST"])
@login_required(role="admin")
def update_user_role(email):
    role = request.form.get("role")
    u = load_users_from_db()
    if email in u:
        save_user_to_db(email, u[email]["password"], role)
        log_admin_action_db(session.get("email"), "Update", email)
        flash(f"Role updated for {email}", "success")
    return redirect(url_for("manage_users"))

@app.route("/update_user_password/<email>", methods=["POST"])
@login_required(role="admin")
def update_user_password(email):
    new_password = (request.form.get("password") or "").strip()
    u = load_users_from_db()
    if email in u and new_password:
        save_user_to_db(email, new_password, u[email]["role"])
        log_admin_action_db(session.get("email"), "Set Password", email)
        flash(f"Password updated for {email}", "success")
    return redirect(url_for("manage_users"))

# ---------------- Health / Utility ----------------
@app.route("/refresh_users", methods=["GET"])
@login_required(role="admin")
def refresh_users():
    try:
        users = zoom_get_users()
        return jsonify({"status":"success","zoom_users":users})
    except Exception as e:
        return jsonify({"status":"error","message":str(e)}), 500
# ---------------- Run ----------------
if __name__ == "__main__":
    # ensure DB created & JSON saved
    init_db()
    save_report_json()
    # run app
    app.run(host="0.0.0.0", port=5000, debug=True)










